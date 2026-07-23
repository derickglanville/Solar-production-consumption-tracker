from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


APPLIANCE_WORKBOOK_PATH = Path(
    r"C:\Software Developement\ChatGPT Codex\Solar Energy - SunRun\Applicances\Appliances.xlsx"
)

APPLIANCE_HEADERS = [
    "Appliance",
    "Room",
    "Typical kW",
    "Typical Hours/Day",
    "Estimated Daily kWh",
    "Estimated Monthly kWh",
    "Recommendation",
    "Estimated Source",
    "Notes",
]

LEGACY_HEADERS_WITH_EXCLUDE = [
    "Exclude from Total",
    "Appliance",
    "Room",
    "Typical kW",
    "Typical Hours/Day",
    "Estimated Daily kWh",
    "Estimated Monthly kWh",
    "Recommendation",
    "Estimated Source",
    "Notes",
]

KNOWN_ROOMS = {
    "basement",
    "family room",
    "kitchen",
    "dining room",
    "living room",
    "garage",
    "master bedroom",
    "office upstairs",
    "office upstaris",
    "theo's room",
    "shanelle's room",
    "unassigned",
}


@dataclass
class ApplianceRecord:
    appliance: str
    room: str
    typical_kw: float
    typical_hours_per_day: float
    estimated_daily_kwh: float
    estimated_monthly_kwh: float
    recommendation: str
    estimated_source: str
    notes: str


def _normalize_name(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace('"', "").split())


def _normalize_room(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


APPLIANCE_ESTIMATES: Dict[str, Dict[str, Any]] = {
    _normalize_name("Clothes Dryer"): {
        "room": "Basement",
        "typical_kw": 5.0,
        "typical_hours_per_day": 0.75,
        "notes": "Typical electric dryer heating load.",
    },
    _normalize_name("Washing machine"): {
        "room": "Basement",
        "typical_kw": 0.5,
        "typical_hours_per_day": 0.75,
        "notes": "Motor and wash-cycle average draw.",
    },
    _normalize_name("Dehumidifyer"): {
        "room": "Basement",
        "typical_kw": 0.6,
        "typical_hours_per_day": 8.0,
        "notes": "Continuous summer moisture-control estimate.",
    },
    _normalize_name("Weil Mclain Boiler"): {
        "room": "Basement",
        "typical_kw": 0.18,
        "typical_hours_per_day": 8.0,
        "notes": "Circulator, blower, and control power only, not fuel heat content.",
    },
    _normalize_name("Hot Water Tank"): {
        "room": "Basement",
        "typical_kw": 0.08,
        "typical_hours_per_day": 10.0,
        "notes": "Indirect hot water tank load represented as added boiler runtime and controls.",
    },
    _normalize_name("Freezer"): {
        "room": "Basement",
        "typical_kw": 0.12,
        "typical_hours_per_day": 10.0,
        "notes": "Compressor-equivalent runtime estimate.",
    },
    _normalize_name("Electric base heater"): {
        "room": "Family Room",
        "typical_kw": 1.5,
        "typical_hours_per_day": 4.0,
        "notes": "Per-heater seasonal estimate; winter usage can be much higher.",
    },
    _normalize_name('55" TV'): {
        "room": "Family Room",
        "typical_kw": 0.09,
        "typical_hours_per_day": 5.0,
        "notes": "Modern LED television estimate.",
    },
    _normalize_name("Refrigderator"): {
        "room": "Kitchen",
        "typical_kw": 0.15,
        "typical_hours_per_day": 10.0,
        "notes": "Compressor-equivalent runtime estimate.",
    },
    _normalize_name("GE Oven"): {
        "room": "Kitchen",
        "typical_kw": 3.5,
        "typical_hours_per_day": 1.5,
        "notes": "Electric oven heating-element estimate; existing workbook note suggested 6-7 kWh per use day.",
    },
    _normalize_name("Dishwasher"): {
        "room": "Kitchen",
        "typical_kw": 1.5,
        "typical_hours_per_day": 1.5,
        "notes": "Heated wash and dry cycle estimate.",
    },
    _normalize_name("Coffee maker"): {
        "room": "Kitchen",
        "typical_kw": 1.0,
        "typical_hours_per_day": 0.25,
        "notes": "Brew cycle estimate.",
    },
    _normalize_name("Rice cooker, blender, toaster"): {
        "room": "Kitchen",
        "typical_kw": 1.2,
        "typical_hours_per_day": 0.3,
        "notes": "Grouped small-kitchen-appliance estimate.",
    },
    _normalize_name("Microware oven 1000 watt"): {
        "room": "Kitchen",
        "typical_kw": 1.0,
        "typical_hours_per_day": 0.25,
        "notes": "Nameplate microwave output rounded to input draw.",
    },
    _normalize_name("Ductless Air AC 220 volt"): {
        "room": "Dining room",
        "typical_kw": 1.2,
        "typical_hours_per_day": 8.0,
        "notes": "Single-zone ductless cooling estimate during active summer use.",
    },
    _normalize_name('65" TV|Living room'): {
        "room": "Living room",
        "typical_kw": 0.12,
        "typical_hours_per_day": 5.0,
        "notes": "Modern large-format LED television estimate.",
    },
    _normalize_name("Garage door opener - 2"): {
        "room": "Garage",
        "typical_kw": 0.5,
        "typical_hours_per_day": 0.1,
        "notes": "Short burst load across two openers.",
    },
    _normalize_name('65" TV|Master bedroom'): {
        "room": "Master bedroom",
        "typical_kw": 0.12,
        "typical_hours_per_day": 5.0,
        "notes": "Modern large-format LED television estimate.",
    },
    _normalize_name("4 Lamps"): {
        "room": "Master bedroom",
        "typical_kw": 0.04,
        "typical_hours_per_day": 5.0,
        "notes": "Assumes four efficient LED lamps.",
    },
    _normalize_name("Portable AC|Master bedroom"): {
        "room": "Master bedroom",
        "typical_kw": 1.1,
        "typical_hours_per_day": 8.0,
        "notes": "Portable air-conditioner summer runtime estimate.",
    },
    _normalize_name("Portable AC|Office upstaris"): {
        "room": "Office upstaris",
        "typical_kw": 1.1,
        "typical_hours_per_day": 8.0,
        "notes": "Portable air-conditioner summer runtime estimate.",
    },
    _normalize_name("Computer monitors - 2"): {
        "room": "Office upstaris",
        "typical_kw": 0.08,
        "typical_hours_per_day": 8.0,
        "notes": "Two monitors combined.",
    },
    _normalize_name("Mini computer Intel NUC"): {
        "room": "Office upstaris",
        "typical_kw": 0.03,
        "typical_hours_per_day": 12.0,
        "notes": "Small-form-factor computer estimate.",
    },
    _normalize_name("Small coffee heater"): {
        "room": "Office upstaris",
        "typical_kw": 0.2,
        "typical_hours_per_day": 4.0,
        "notes": "Mug warmer / small plate heater estimate.",
    },
    _normalize_name('55" TV not used'): {
        "room": "Theo's room",
        "typical_kw": 0.08,
        "typical_hours_per_day": 0.0,
        "notes": "Set as effectively off based on workbook label.",
    },
    _normalize_name("Desktop computer - runs oftn"): {
        "room": "Theo's room",
        "typical_kw": 0.25,
        "typical_hours_per_day": 8.0,
        "notes": "Desktop system estimate for regular active use.",
    },
    _normalize_name("Computer monitor"): {
        "room": "Theo's room",
        "typical_kw": 0.04,
        "typical_hours_per_day": 8.0,
        "notes": "Single monitor estimate.",
    },
    _normalize_name("Lamps -2|Theo's room"): {
        "room": "Theo's room",
        "typical_kw": 0.02,
        "typical_hours_per_day": 5.0,
        "notes": "Two efficient LED lamps.",
    },
    _normalize_name("Portable AC|Theo's room"): {
        "room": "Theo's room",
        "typical_kw": 1.1,
        "typical_hours_per_day": 8.0,
        "notes": "Portable air-conditioner summer runtime estimate.",
    },
    _normalize_name("Lamps -2|Shanelle's room"): {
        "room": "Shanelle's room",
        "typical_kw": 0.02,
        "typical_hours_per_day": 5.0,
        "notes": "Two efficient LED lamps.",
    },
}


def _safe_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, str):
        cleaned = value.strip().replace("kWh", "").replace("KW", "").replace("kW", "")
        cleaned = cleaned.replace("\uFFFD", "").replace(" ", "")
        if "-" in cleaned and cleaned.replace(".", "", 2).replace("-", "", 1).isdigit():
            parts = [part for part in cleaned.split("-") if part]
            if len(parts) == 2:
                try:
                    return round((float(parts[0]) + float(parts[1])) / 2.0, 3)
                except ValueError:
                    return default
        try:
            return float(cleaned)
        except ValueError:
            return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _round_energy(value: float) -> float:
    return round(float(value or 0.0), 2)


def _record_to_dict(record: ApplianceRecord) -> Dict[str, Any]:
    return asdict(record)


def _build_recommendation(appliance_name: str, typical_kw: float, hours_per_day: float) -> str:
    normalized = _normalize_name(appliance_name)
    monthly_kwh = typical_kw * hours_per_day * 30.0

    if "hot water tank" in normalized:
        return "Indirect load through the boiler. Lower standby losses with pipe insulation, tank temperature tuning, and lower hot-water waste."
    if "portable ac" in normalized:
        return "High summer load. Seal the room, clean the filter, and consider a higher-efficiency cooling option."
    if "ductless air ac" in normalized:
        return "Major cooling load. Clean filters and use setback temperatures to reduce runtime."
    if "electric base heater" in normalized:
        return "Resistance heat is expensive. Reduce runtime where possible and consider a heat-pump upgrade."
    if "clothes dryer" in normalized:
        return "Run full loads, clean the vent, and use lower-heat cycles when practical."
    if "dehumid" in normalized:
        return "Long runtime adds up. Raise the humidity setpoint a bit and keep the area sealed."
    if "refrig" in normalized or "freezer" in normalized:
        return "Check seals and keep coils clean so the compressor does not run longer than needed."
    if "oven" in normalized or "dishwasher" in normalized:
        return "Batch cooking or washing helps. Eco modes can reduce energy use."
    if "tv" in normalized or "monitor" in normalized or "computer" in normalized:
        return "Enable sleep timers and aggressive power-saving settings."
    if "lamp" in normalized:
        return "Low load if LED. Swap any remaining non-LED bulbs first."
    if monthly_kwh >= 112:
        return "This is a major estimated energy user. Reducing runtime will likely matter more than small efficiency tweaks."
    if monthly_kwh >= 19:
        return "Worth monitoring. A small cut in daily runtime could noticeably reduce consumption."
    return "Lower estimated load. Focus first on higher-kWh appliances unless usage changes a lot."


def _build_default_estimate(appliance_name: str, room: str = "") -> Dict[str, Any]:
    normalized = _normalize_name(appliance_name)
    room_normalized = _normalize_room(room)
    room_key = _normalize_name(f"{appliance_name}|{room}")
    if room_key in APPLIANCE_ESTIMATES:
        estimate = APPLIANCE_ESTIMATES[room_key]
        return {
            "typical_kw": float(estimate["typical_kw"]),
            "typical_hours_per_day": float(estimate["typical_hours_per_day"]),
            "estimated_source": "Estimated baseline",
            "notes": str(estimate.get("notes", "")),
            "room": estimate.get("room") or room,
        }
    if normalized in APPLIANCE_ESTIMATES:
        estimate = APPLIANCE_ESTIMATES[normalized]
        return {
            "typical_kw": float(estimate["typical_kw"]),
            "typical_hours_per_day": float(estimate["typical_hours_per_day"]),
            "estimated_source": "Estimated baseline",
            "notes": str(estimate.get("notes", "")),
            "room": estimate.get("room") or room,
        }
    return {
        "typical_kw": 0.1,
        "typical_hours_per_day": 1.0,
        "estimated_source": "Fallback estimate",
        "notes": "Fallback estimate. Replace with a device-specific value if you know the nameplate wattage or runtime.",
        "room": room_normalized.title() if room_normalized else room,
    }


def _infer_appliance_from_shifted_row(
    room: str,
    typical_kw: float,
    typical_hours_per_day: float,
    notes: str,
) -> str:
    notes_normalized = str(notes or "").strip().lower()
    room_normalized = _normalize_room(room)
    best_match: Optional[str] = None

    for key, estimate in APPLIANCE_ESTIMATES.items():
        estimate_room = _normalize_room(str(estimate.get("room", "")))
        estimate_notes = str(estimate.get("notes", "")).strip().lower()
        if room_normalized and estimate_room and room_normalized != estimate_room:
            continue
        if abs(float(estimate["typical_kw"]) - typical_kw) > 0.01:
            continue
        if abs(float(estimate["typical_hours_per_day"]) - typical_hours_per_day) > 0.01:
            continue
        if estimate_notes and estimate_notes[:24] and estimate_notes[:24] in notes_normalized:
            return key.split("|")[0]
        best_match = key.split("|")[0]

    return best_match or "Recovered Appliance"


def _row_is_empty(values: List[Any]) -> bool:
    return not any(str(value).strip() for value in values if value is not None)


def _parse_row_by_header(values: List[Any], headers: List[str]) -> ApplianceRecord:
    header_map = {str(header).strip(): index for index, header in enumerate(headers)}
    appliance = str(values[header_map["Appliance"]] or "").strip()
    room = str(values[header_map["Room"]] or "").strip()
    default_estimate = _build_default_estimate(appliance, room)
    typical_kw = _safe_float(values[header_map["Typical kW"]], default_estimate["typical_kw"])
    typical_hours_per_day = _safe_float(
        values[header_map["Typical Hours/Day"]], default_estimate["typical_hours_per_day"]
    )
    estimated_daily_kwh = _round_energy(typical_kw * typical_hours_per_day)
    estimated_monthly_kwh = _round_energy(estimated_daily_kwh * 30.0)
    recommendation = str(
        values[header_map.get("Recommendation", -1)] if "Recommendation" in header_map else ""
    ).strip() or _build_recommendation(appliance, typical_kw, typical_hours_per_day)
    estimated_source = str(
        values[header_map.get("Estimated Source", -1)] if "Estimated Source" in header_map else ""
    ).strip() or default_estimate["estimated_source"]
    notes = str(values[header_map.get("Notes", -1)] if "Notes" in header_map else "").strip() or default_estimate["notes"]
    return ApplianceRecord(
        appliance=appliance,
        room=room,
        typical_kw=typical_kw,
        typical_hours_per_day=typical_hours_per_day,
        estimated_daily_kwh=estimated_daily_kwh,
        estimated_monthly_kwh=estimated_monthly_kwh,
        recommendation=recommendation,
        estimated_source=estimated_source,
        notes=notes,
    )


def _parse_shifted_row(values: List[Any]) -> ApplianceRecord:
    room = str(values[0] or "").strip()
    typical_kw = _safe_float(values[1], 0.1)
    typical_hours_per_day = _safe_float(values[2], 1.0)
    source_or_notes = str(values[7] if len(values) > 7 else "").strip()
    notes = str(values[8] if len(values) > 8 else "").strip()
    appliance = _infer_appliance_from_shifted_row(room, typical_kw, typical_hours_per_day, notes or source_or_notes)
    default_estimate = _build_default_estimate(appliance, room)
    estimated_daily_kwh = _round_energy(typical_kw * typical_hours_per_day)
    estimated_monthly_kwh = _round_energy(estimated_daily_kwh * 30.0)
    return ApplianceRecord(
        appliance=appliance,
        room=room,
        typical_kw=typical_kw,
        typical_hours_per_day=typical_hours_per_day,
        estimated_daily_kwh=estimated_daily_kwh,
        estimated_monthly_kwh=estimated_monthly_kwh,
        recommendation=_build_recommendation(appliance, typical_kw, typical_hours_per_day),
        estimated_source=str(values[7] if len(values) > 7 else "").strip() or default_estimate["estimated_source"],
        notes=notes or default_estimate["notes"],
    )


def _row_looks_shifted(values: List[Any]) -> bool:
    if len(values) < 5:
        return False
    first = _normalize_room(values[0])
    second = _safe_float(values[1], -999)
    third = _safe_float(values[2], -999)
    fourth = _safe_float(values[3], -999)
    fifth = _safe_float(values[4], -999)
    return (
        first in KNOWN_ROOMS
        and second >= 0
        and third >= 0
        and fourth >= 0
        and fifth >= 0
    )


def _parse_any_row(values: List[Any], headers: List[str]) -> ApplianceRecord:
    if headers == APPLIANCE_HEADERS:
        if _row_looks_shifted(values):
            return _parse_shifted_row(values)
        return _parse_row_by_header(values, headers)
    if headers == LEGACY_HEADERS_WITH_EXCLUDE:
        legacy_values = values[1:]
        return _parse_row_by_header(legacy_values, APPLIANCE_HEADERS)
    if _row_looks_shifted(values):
        return _parse_shifted_row(values)
    return _parse_row_by_header(values, APPLIANCE_HEADERS)


def _rewrite_workbook(records: List[ApplianceRecord]):
    workbook = load_workbook(APPLIANCE_WORKBOOK_PATH)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(APPLIANCE_HEADERS)
    for record in records:
        worksheet.append(
            [
                record.appliance,
                record.room,
                _round_energy(record.typical_kw),
                _round_energy(record.typical_hours_per_day),
                _round_energy(record.estimated_daily_kwh),
                _round_energy(record.estimated_monthly_kwh),
                record.recommendation,
                record.estimated_source,
                record.notes,
            ]
        )
    workbook.save(APPLIANCE_WORKBOOK_PATH)


def _ensure_default_hot_water_tank(records: List[ApplianceRecord]) -> List[ApplianceRecord]:
    if any(_normalize_name(record.appliance) == _normalize_name("Hot Water Tank") for record in records):
        return records
    default_estimate = _build_default_estimate("Hot Water Tank", "Basement")
    typical_kw = default_estimate["typical_kw"]
    typical_hours_per_day = default_estimate["typical_hours_per_day"]
    records.append(
        ApplianceRecord(
            appliance="Hot Water Tank",
            room="Basement",
            typical_kw=typical_kw,
            typical_hours_per_day=typical_hours_per_day,
            estimated_daily_kwh=_round_energy(typical_kw * typical_hours_per_day),
            estimated_monthly_kwh=_round_energy(typical_kw * typical_hours_per_day * 30.0),
            recommendation=_build_recommendation("Hot Water Tank", typical_kw, typical_hours_per_day),
            estimated_source=default_estimate["estimated_source"],
            notes=default_estimate["notes"],
        )
    )
    return records


def _load_and_normalize_records() -> List[ApplianceRecord]:
    workbook = load_workbook(APPLIANCE_WORKBOOK_PATH)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        records: List[ApplianceRecord] = _ensure_default_hot_water_tank([])
        _rewrite_workbook(records)
        return records

    headers = [str(value or "").strip() for value in rows[0]]
    source_rows = [list(row) for row in rows[1:] if not _row_is_empty(list(row))]

    parsed_records = [_parse_any_row(values, headers) for values in source_rows]
    parsed_records = _ensure_default_hot_water_tank(parsed_records)
    parsed_records.sort(key=lambda record: record.estimated_monthly_kwh, reverse=True)

    if headers != APPLIANCE_HEADERS or any(_row_looks_shifted(values) for values in source_rows):
        _rewrite_workbook(parsed_records)

    return parsed_records


def load_appliance_summary() -> Dict[str, Any]:
    records = _load_and_normalize_records()
    total_monthly_kwh = _round_energy(sum(row.estimated_monthly_kwh for row in records))
    total_daily_kwh = _round_energy(sum(row.estimated_daily_kwh for row in records))
    highest_kw_record = max(records, key=lambda row: row.typical_kw, default=None)
    highest_monthly_record = records[0] if records else None
    room_totals: Dict[str, float] = {}
    for record in records:
        room_totals[record.room] = room_totals.get(record.room, 0.0) + record.estimated_monthly_kwh
    top_rooms = sorted(room_totals.items(), key=lambda item: item[1], reverse=True)[:5]

    return {
        "available": APPLIANCE_WORKBOOK_PATH.exists(),
        "workbook_name": APPLIANCE_WORKBOOK_PATH.name,
        "workbook_path": str(APPLIANCE_WORKBOOK_PATH),
        "record_count": len(records),
        "total_monthly_kwh": total_monthly_kwh,
        "total_daily_kwh": total_daily_kwh,
        "highest_kw_record": _record_to_dict(highest_kw_record) if highest_kw_record else None,
        "highest_monthly_record": _record_to_dict(highest_monthly_record) if highest_monthly_record else None,
        "top_rooms": [
            {"room": room, "estimated_monthly_kwh": _round_energy(total)}
            for room, total in top_rooms
        ],
        "records": [_record_to_dict(record) for record in records],
    }


def save_appliance_records(payload_records: List[Dict[str, Any]]) -> Dict[str, Any]:
    records: List[ApplianceRecord] = []
    for item in payload_records:
        appliance = str(item.get("appliance", "")).strip()
        room = str(item.get("room", "")).strip()
        default_estimate = _build_default_estimate(appliance, room)
        typical_kw = _safe_float(item.get("typical_kw"), default_estimate["typical_kw"])
        typical_hours_per_day = _safe_float(
            item.get("typical_hours_per_day"), default_estimate["typical_hours_per_day"]
        )
        estimated_daily_kwh = _round_energy(typical_kw * typical_hours_per_day)
        estimated_monthly_kwh = _round_energy(estimated_daily_kwh * 30.0)
        recommendation = str(
            item.get("recommendation") or _build_recommendation(appliance, typical_kw, typical_hours_per_day)
        ).strip()
        estimated_source = str(item.get("estimated_source") or "User updated").strip()
        notes = str(item.get("notes") or default_estimate["notes"]).strip()

        records.append(
            ApplianceRecord(
                appliance=appliance,
                room=room,
                typical_kw=typical_kw,
                typical_hours_per_day=typical_hours_per_day,
                estimated_daily_kwh=estimated_daily_kwh,
                estimated_monthly_kwh=estimated_monthly_kwh,
                recommendation=recommendation,
                estimated_source=estimated_source,
                notes=notes,
            )
        )

    records = _ensure_default_hot_water_tank(records)
    records.sort(key=lambda record: record.estimated_monthly_kwh, reverse=True)
    _rewrite_workbook(records)
    return load_appliance_summary()
